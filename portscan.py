import openpyxl
import shodan

# Shodan API Key (replace with your own API key)
SHODAN_API_KEY = "XX"

# Function to query Shodan and get open ports for an IP address
def query_shodan(ip_address, api_key):
    try:
        api = shodan.Shodan(api_key)
        result = api.host(ip_address)
        return result['ports']
    except shodan.APIError as e:
        print(f"Error querying Shodan for {ip_address}: {e}")
        return []

# Load the Excel spreadsheet
input_filename = "testip.xlsx"
output_filename = "output.xlsx"
workbook = openpyxl.load_workbook(input_filename)
sheet = workbook.active

# Iterate through rows in the spreadsheet, assuming IP addresses are in column A
for row_number, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
    ip_address = row[0]
    open_ports = query_shodan(ip_address, SHODAN_API_KEY)

    # Write open ports back to the spreadsheet in column B
    sheet.cell(row=row_number, column=2, value=','.join(map(str, open_ports)))

# Save the updated spreadsheet
workbook.save(output_filename)

print("Open ports have been written to the spreadsheet.")
